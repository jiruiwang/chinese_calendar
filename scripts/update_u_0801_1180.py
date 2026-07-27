from pathlib import Path
from openpyxl import load_workbook
import json,re
ROOT=Path(__file__).resolve().parents[1]
SRC=ROOT/'source'/'U_0801_1180.xlsx'

def norm(v):
    if v is None:return ''
    return str(v).strip().replace('歷','曆').replace('历','曆').replace('錄','録').replace('录','録').replace('吳','吴').replace('黃','黄').replace('陈','陳').replace('后','後')

def load(y):
    p=ROOT/'data'/f'{y}.js';s=p.read_text(encoding='utf-8')
    m=re.search(r'window\.CalendarData\s*=\s*(\{.*?\});\s*\n\nwindow\.CalendarData\.dayIndex',s,re.S)
    return p,json.loads(m.group(1))

def save(p,d):
    p.write_text('window.CalendarData = '+json.dumps(d,ensure_ascii=False,separators=(',',':'))+';\n\nwindow.CalendarData.dayIndex = Object.fromEntries(\n  window.CalendarData.months.flatMap(month => month.days.map(day => [`${day.key}|${day.western.date}`, day]))\n);\n',encoding='utf-8')

wb=load_workbook(SRC,read_only=True,data_only=True);ws=wb.active
updates={}
for date,era in ws.iter_rows(min_row=3,values_only=True):
    date=norm(date);era=norm(era)
    if date and 801<=int(date[:4])<=1180: updates[date]=era
assert len(updates)>130000, len(updates)

changed=0
for y in range(801,1181):
    p,d=load(y)
    first_key=None
    for m in d['months']:
        for day in m['days']:
            if first_key is None:first_key=day['chinese']['date']
            key=day['chinese']['date']
            for rec in day.get('orthodoxies',[]):
                if rec.get('group')=='正朔一' and key in updates:
                    if rec.get('eraYear','')!=updates[key]: changed+=1
                    rec['eraYear']=updates[key]
    # Rebuild title eras from all current daily orthodoxies, preserving first-seen state/order.
    state_order=[]; entries={}; seen=set()
    for m in d['months']:
        for day in m['days']:
            key=day['chinese']['date']
            for rec in day.get('orthodoxies',[]):
                st=norm(rec.get('state')); era=norm(rec.get('eraYear'))
                if not st or not era: continue
                if st not in entries: entries[st]=[];state_order.append(st)
                pair=(st,era)
                if pair in seen: continue
                seen.add(pair)
                label=era
                if key!=first_key:
                    md=(norm(rec.get('month'))+norm(rec.get('day'))).strip()
                    if md: label=f'{era}（{md}改元）'
                entries[st].append(label)
    d['title']['eras']=[{'state':st,'entries':entries[st]} for st in state_order]
    save(p,d)

# Patch index years for 801-1180 from rebuilt title.
ip=ROOT/'data'/'index-data.js';txt=ip.read_text(encoding='utf-8')
obj=json.loads(re.search(r'window\.CalendarIndexData\s*=\s*(\{.*\});',txt,re.S).group(1))
mp={x['year']:x for x in obj['years']}
for y in range(801,1181):
    _,d=load(y);t=d['title'];details=[]
    for g in t.get('eras',[]):
        for full in g.get('entries',[]):
            base=re.sub(r'（.*?）|\(.*?\)','',full).strip()
            details.append({'state':g['state'],'text':base,'fullText':full})
    mp[y]={'year':y,'year4':f'{y:04d}','href':f'years/{y}.html','ganzhi':t['ganzhi'],'solarYear':t['solarNewYear'].split('-')[0],'eras':[x['text'] for x in details],'states':list(dict.fromkeys(x['state'] for x in details)),'eraDetails':details}
obj['years']=[mp[y] for y in sorted(mp)]
ip.write_text('window.CalendarIndexData = '+json.dumps(obj,ensure_ascii=False,separators=(',',':'))+';\n',encoding='utf-8')
print('rows',len(updates),'changed',changed)
