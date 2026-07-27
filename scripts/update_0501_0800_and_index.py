from pathlib import Path
import json,re
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]
SRC=ROOT/'source'/'川象曆_0501-0800_日表.xlsx'

def c(v):
    if v is None:return ''
    if isinstance(v,float) and v.is_integer(): v=int(v)
    return str(v).strip().replace('歷','曆').replace('历','曆').replace('錄','録').replace('录','録').replace('吳','吴').replace('黃','黄').replace('陈','陳').replace('后','後')
def cd(v):
    s=c(v);m=re.match(r'^(\d+)-(\d{2}[CL])-(\d{2})$',s)
    return f'{int(m.group(1)):04d}-{m.group(2)}-{m.group(3)}' if m else s
def wd(v):
    s=c(v);m=re.match(r'^(B?)(\d+)-(\d{2})-(\d{2})$',s)
    return f'{m.group(1)}{int(m.group(2)):04d}-{m.group(3)}-{m.group(4)}' if m else s
def mn(code):
    names={1:'正月',2:'二月',3:'三月',4:'四月',5:'五月',6:'六月',7:'七月',8:'八月',9:'九月',10:'十月',11:'十一月',12:'十二月'}
    return ('閏' if code.endswith('L') else '')+names[int(code[:2])]
def group(row,start,label):
    a=[c(row[start+i]) if start+i<len(row) else '' for i in range(6)]
    return None if not any(a) else {'group':label,'state':a[0],'ruler':a[1],'eraYear':a[2],'month':a[3],'day':a[4],'calendar':a[5]}

def parse_year_eras(vals):
    eras=[]
    for val in vals:
        for line in c(val).splitlines():
            line=line.strip()
            if not line: continue
            parts=line.split(None,1)
            state=parts[0]; entry=parts[1] if len(parts)>1 else ''
            if eras and eras[-1]['state']==state: eras[-1]['entries'].append(entry)
            else: eras.append({'state':state,'entries':[entry] if entry else []})
    return eras

wb=load_workbook(SRC,read_only=True,data_only=True)
meta={}
for r in wb['年'].iter_rows(min_row=2,values_only=True):
    if r[0] is None: continue
    y=int(float(r[0]))
    if 501<=y<=800:
        meta[y]={'year':f'{y:04d}','displayYear':y,'ganzhi':c(r[1]),'monthCount':int(float(r[2])),'dayCount':int(float(r[3])),'yuanriGanzhi':c(r[4]),'solarNewYear':wd(r[5]),'winterSolstice':cd(r[6]),'eras':parse_year_eras(r[7:10])}
months={y:[] for y in range(501,801)}
last_month={y:None for y in range(501,801)}
for r in wb['日'].iter_rows(min_row=3,values_only=True):
    key=cd(r[0])
    if not key: continue
    y=int(key[:4])
    if not 501<=y<=800: continue
    mc=c(r[4]); western=wd(r[2]); wp=western.split('-')
    orth=[x for x in (group(r,18,'正朔一'),group(r,24,'正朔二'),group(r,30,'正朔三')) if x]
    day={'key':key,'ganzhi':c(r[1]),'weekday':c(r[9]),'chinese':{'date':key,'year':f'{y:04d}','month':mc,'day':c(r[5]).zfill(2),'calendar':c(r[16])},'western':{'date':western,'year':wp[0],'month':wp[1],'day':wp[2],'calendar':c(r[17])},'astronomy':{'syzygy':c(r[10]),'meanSolarTerm':c(r[11]),'moonPhase':c(r[12]),'trueSolarTerm':c(r[13]),'solarEclipse':c(r[14]),'lunarEclipse':c(r[15])},'orthodoxies':orth}
    # Strict source row order. A new month block begins whenever code changes OR day resets to 01.
    if not months[y] or months[y][-1]['id']!=mc or c(r[5]).zfill(2)=='01' and months[y][-1]['days']:
        months[y].append({'id':mc,'name':mn(mc),'isLeap':mc.endswith('L'),'daysInMonth':0,'days':[]})
    months[y][-1]['days'].append(day);months[y][-1]['daysInMonth']+=1
for y in range(501,801):
    if y not in meta: raise RuntimeError(f'missing year meta {y}')
    payload={'year':f'{y:04d}','displayYear':y,'title':meta[y],'months':months[y]}
    text='window.CalendarData = '+json.dumps(payload,ensure_ascii=False,separators=(',',':'))+';\n\nwindow.CalendarData.dayIndex = Object.fromEntries(\n  window.CalendarData.months.flatMap(month => month.days.map(day => [`${day.key}|${day.western.date}`, day]))\n);\n'
    (ROOT/'data'/f'{y}.js').write_text(text,encoding='utf-8')

# Parse all yearly JS and rebuild index records.
def load_data(y):
    s=(ROOT/'data'/f'{y}.js').read_text(encoding='utf-8')
    m=re.search(r'window\.CalendarData\s*=\s*(\{.*?\});\s*\n\nwindow\.CalendarData\.dayIndex',s,re.S)
    if not m: raise RuntimeError(f'cannot parse data/{y}.js')
    return json.loads(m.group(1))

years=[]; datasets={}
for y in range(200,2132):
    d=load_data(y);datasets[y]=d;t=d['title']
    details=[]
    for eg in t.get('eras',[]):
        for entry in eg.get('entries',[]):
            short=re.sub(r'（.*?）|\(.*?\)','',entry).strip()
            details.append({'state':eg.get('state',''),'text':short,'fullText':entry})
    solar=t.get('solarNewYear','').split('-')[0]
    years.append({'year':y,'year4':f'{y:04d}','href':f'years/{y}.html','ganzhi':t.get('ganzhi',''),'solarYear':solar,'eras':[x['text'] for x in details],'states':list(dict.fromkeys(x['state'] for x in details)),'eraDetails':details})

# Preserve dynasty catalogue and ranges, correcting East Han through 0440.
oldtxt=(ROOT/'data'/'index-data.js').read_text(encoding='utf-8')
old=json.loads(re.search(r'window\.CalendarIndexData\s*=\s*(\{.*\});',oldtxt,re.S).group(1))
specs=[]
for d in old['dynasties']:
    start=min(d['years']) if d.get('years') else int(re.search(r'(\d{4})',d['displayRange']).group(1))
    end=max(d['years']) if d.get('years') else int(re.findall(r'(\d{4})',d['displayRange'])[-1])
    if d['id']=='東漢': start,end=245,440
    specs.append((d['id'],d['label'],d['states'],start,end))

def matches_state(state,states): return state in states

dynasties=[]
for did,label,states,start,end in specs:
    dys=[]; ruler_years={}
    for y in range(start,end+1):
        d=datasets[y]; found=False
        for mo in d['months']:
            for day in mo['days']:
                for o in day.get('orthodoxies',[]):
                    if matches_state(o.get('state',''),states):
                        found=True
                        rn=o.get('ruler','').strip()
                        if rn: ruler_years.setdefault(rn,set()).add(y)
                if found and all(False for _ in []): pass
        if found: dys.append(y)
    rulers=[]
    for name,ys in ruler_years.items():
        sy=sorted(ys); rulers.append({'name':name,'years':sy,'range':f'{sy[0]:04d}年至{sy[-1]:04d}年','count':len(sy)})
    rulers.sort(key=lambda r:(r['years'][0],r['name']))
    if dys:
        dynasties.append({'id':did,'label':label,'states':states,'years':dys,'displayRange':f'{dys[0]:04d}年至{dys[-1]:04d}年','displayCount':f'凡{len(dys)}年','allRulersLabel':f'全部國君（{len(rulers)}）','rulers':rulers})
obj={'years':years,'dynasties':dynasties}
(ROOT/'data'/'index-data.js').write_text('window.CalendarIndexData = '+json.dumps(obj,ensure_ascii=False,separators=(',',':'))+';\n',encoding='utf-8')
print('updated 0501-0800 and rebuilt index data')
