#!/usr/bin/env python3
"""Rebuild data/0401.js through data/0500.js from the two source XLSX files.
Run from the project root: python scripts/build-calendar-data.py
"""
from pathlib import Path
from collections import OrderedDict
import json, re
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]
DAY=ROOT/'source'/'川象曆_0401-0500_日表.xlsx'
YEAR=ROOT/'source'/'川象曆_0401-0500_年表.xlsx'
YEARS=range(401,501)
def c(v):
    return '' if v is None else str(int(v) if isinstance(v,float) and v.is_integer() else v).strip().replace('歷','曆').replace('历','曆').replace('吳','吴')
def cd(v):
    s=c(v); m=re.match(r'^(\d+)-(\d{2}[CL])-(\d{2})$',s); return f'{int(m.group(1)):04d}-{m.group(2)}-{m.group(3)}' if m else s
def wd(v):
    s=c(v); m=re.match(r'^(B?)(\d+)-(\d{2})-(\d{2})$',s); return f'{m.group(1)}{int(m.group(2)):04d}-{m.group(3)}-{m.group(4)}' if m else s
def mn(code):
    names={1:'正月',2:'二月',3:'三月',4:'四月',5:'五月',6:'六月',7:'七月',8:'八月',9:'九月',10:'十月',11:'十一月',12:'十二月'}
    return ('閏' if code.endswith('L') else '')+names[int(code[:2])]
def group(row,start,label):
    a=[c(row[start+i]) if start+i<len(row) else '' for i in range(6)]
    return None if not any(a) else {'group':label,'state':a[0],'ruler':a[1],'eraYear':a[2],'month':a[3],'day':a[4],'calendar':a[5]}
ywb=load_workbook(YEAR,read_only=True,data_only=True); yws=ywb['年']; meta={}
for r in yws.iter_rows(min_row=2,values_only=True):
    if r[0] is None: continue
    y=int(float(r[0]));
    if y not in YEARS: continue
    vals={'漢':r[7],'魏':r[8],'晉':r[9],'吴':r[10]}; order=['漢','魏','晉','吴'] if y<=440 else ['魏','晉','漢','吴']; eras=[]
    for state in order:
        entries=[c(x) for x in c(vals[state]).splitlines() if c(x)]
        if entries: eras.append({'state':state,'entries':entries})
    meta[y]={'year':f'{y:04d}','displayYear':y,'ganzhi':c(r[1]),'monthCount':int(r[2]),'dayCount':int(r[3]),'yuanriGanzhi':c(r[4]),'solarNewYear':wd(r[5]),'winterSolstice':cd(r[6]),'eras':eras}
months={y:OrderedDict() for y in YEARS}; dwb=load_workbook(DAY,read_only=True,data_only=True); dws=dwb['日']
for r in dws.iter_rows(min_row=3,values_only=True):
    key=cd(r[0]);
    if not key: continue
    y=int(key[:4]);
    if y not in months: continue
    mc=c(r[4]); western=wd(r[2]); wp=western.split('-'); orthodoxies=[x for x in [group(r,18,'正朔一'),group(r,24,'正朔二'),group(r,30,'正朔三')] if x]
    day={'key':key,'ganzhi':c(r[1]),'weekday':c(r[9]),'chinese':{'date':key,'year':f'{y:04d}','month':mc,'day':c(r[5]).zfill(2),'calendar':c(r[16])},'western':{'date':western,'year':wp[0],'month':wp[1],'day':wp[2],'calendar':c(r[17])},'astronomy':{'syzygy':c(r[10]),'meanSolarTerm':c(r[11]),'moonPhase':c(r[12]),'trueSolarTerm':c(r[13]),'solarEclipse':c(r[14]),'lunarEclipse':c(r[15])},'orthodoxies':orthodoxies}
    if mc not in months[y]: months[y][mc]={'id':mc,'name':mn(mc),'isLeap':mc.endswith('L'),'daysInMonth':0,'days':[]}
    months[y][mc]['days'].append(day); months[y][mc]['daysInMonth']+=1
for y in YEARS:
    payload={'year':f'{y:04d}','displayYear':y,'title':meta[y],'months':list(months[y].values())}
    text='window.CalendarData = '+json.dumps(payload,ensure_ascii=False,separators=(',',':'))+';\n\nwindow.CalendarData.dayIndex = Object.fromEntries(\n  window.CalendarData.months.flatMap(month => month.days.map(day => [day.key, day]))\n);\n'
    (ROOT/'data'/f'{y}.js').write_text(text,encoding='utf-8')
    print(f'generated data/{y}.js')
