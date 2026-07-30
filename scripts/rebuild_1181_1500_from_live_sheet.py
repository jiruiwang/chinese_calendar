from __future__ import annotations
from pathlib import Path
from collections import defaultdict
import json, re, sys
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / 'source' / '川象曆_1181-1500_live.xlsx'
START, END = 1181, 1500


def s(v):
    if v is None: return ''
    if isinstance(v, float) and v.is_integer(): return str(int(v))
    return str(v).strip()


def parse_era_cell(value):
    entries=[]
    for line in s(value).splitlines():
        line=line.strip()
        if not line: continue
        if '：' in line:
            era, source=line.split('：',1)
        elif ':' in line:
            era, source=line.split(':',1)
        else:
            era, source=line, ''
        # 0501—2131年号标题不显示改元/改年括注。
        era=re.sub(r'[（(][^）)]*(?:改元|改年)[^）)]*[）)]', '', era).strip()
        notes=[x.strip() for x in source.splitlines() if x.strip()]
        item={'era':era,'source':source.strip(),'notes':notes}
        entries.append(item)
    return entries


def compact_entry(item):
    # Keep object form so title source remains available to the UI and search index.
    return item


def load_index():
    p=ROOT/'data'/'index-data.js'
    text=p.read_text(encoding='utf-8')
    m=re.search(r'window\.CalendarIndexData\s*=\s*(\{.*\});\s*$',text,re.S)
    if not m: raise RuntimeError('index-data.js parse failed')
    return p,json.loads(m.group(1))


def dump_js(path,data):
    text='window.CalendarData = '+json.dumps(data,ensure_ascii=False,separators=(',',':'))+';\n\n'
    text+='window.CalendarData.dayIndex = Object.fromEntries(\n  window.CalendarData.months.flatMap(month => month.days.map(day => [`${day.key}|${day.western.date}`, day]))\n);\n'
    path.write_text(text,encoding='utf-8')


def month_name(code, recs):
    for r in recs:
        if r.get('month'): return r['month']
    n=int(re.match(r'\d+',code).group()) if re.match(r'\d+',code) else 0
    base={1:'正月',2:'二月',3:'三月',4:'四月',5:'五月',6:'六月',7:'七月',8:'八月',9:'九月',10:'十月',11:'十一月',12:'十二月'}.get(n,code)
    return ('閏'+base) if code.endswith('L') else base


def main():
    wb=load_workbook(XLSX,read_only=True,data_only=True)
    ws_day=wb['日']; ws_year=wb['年']

    # Year page
    year_rows={}
    it=ws_year.iter_rows(values_only=True)
    yheaders=[s(x) for x in next(it)]
    for row in it:
        y=s(row[0])
        if not y: continue
        yi=int(float(y))
        if START<=yi<=END:
            year_rows[yi]={yheaders[i]:row[i] if i<len(row) else None for i in range(len(yheaders))}

    # Day page: two header rows, preserving source order.
    it=ws_day.iter_rows(values_only=True)
    h1=[s(x) for x in next(it)]; h2=[s(x) for x in next(it)]
    by_year=defaultdict(list)
    total=0
    for row in it:
        key=s(row[0] if row else '')
        if not key: continue
        year=s(row[3] if len(row)>3 else '')
        if not year: continue
        yi=int(float(year))
        if not (START<=yi<=END): continue
        vals=[s(row[i] if i<len(row) else '') for i in range(40)]
        recs=[]
        for base,group in ((18,'正朔一'),(24,'正朔二'),(30,'正朔三')):
            fields=vals[base:base+6]
            if any(fields):
                recs.append({'group':group,'state':fields[0],'ruler':fields[1],'eraYear':fields[2],
                             'month':fields[3],'day':fields[4],'calendar':fields[5]})
        day={
          'key':vals[0], 'ganzhi':vals[1], 'weekday':vals[9],
          'chinese':{'date':vals[0],'year':vals[3],'month':vals[4],'day':vals[5],'calendar':vals[16]},
          'western':{'date':vals[2],'year':vals[6],'month':vals[7],'day':vals[8],'calendar':vals[17]},
          'astronomy':{'syzygy':vals[10],'meanSolarTerm':vals[11],'moonPhase':vals[12],
                       'trueSolarTerm':vals[13],'solarEclipse':vals[14],'lunarEclipse':vals[15]},
          'orthodoxies':recs,
          'checks':{'改曆':vals[36],'新君':vals[37],'改元':vals[38]},
          'source':vals[39]
        }
        by_year[yi].append(day); total+=1

    missing=sorted(set(range(START,END+1))-set(year_rows))
    if missing: raise RuntimeError(f'missing year rows: {missing[:10]}')
    missing_days=sorted(y for y in range(START,END+1) if not by_year[y])
    if missing_days: raise RuntimeError(f'missing day rows: {missing_days[:10]}')

    generated={}
    for y in range(START,END+1):
        yr=year_rows[y]; days=by_year[y]
        groups=[]
        for state in ('宋','遼','金','元'):
            entries=parse_era_cell(yr.get(state))
            if entries: groups.append({'state':state,'entries':[compact_entry(x) for x in entries]})
        months=[]; cur=None
        for d in days:
            code=d['chinese']['month']; dn=d['chinese']['day']
            # Continuous month segments: a new day 01 always starts a new month, even when code repeats.
            if cur is None or code!=cur['id'] or dn=='01':
                cur={'id':code,'name':month_name(code,d['orthodoxies']),
                     'isLeap':code.endswith('L'),'daysInMonth':0,'days':[]}
                months.append(cur)
            cur['days'].append(d); cur['daysInMonth']+=1
        expected=int(float(s(yr.get('日數'))))
        if len(days)!=expected:
            raise RuntimeError(f'{y}: days {len(days)} != year page {expected}')
        expected_months=int(float(s(yr.get('月數'))))
        if len(months)!=expected_months:
            raise RuntimeError(f'{y}: months {len(months)} != year page {expected_months}')
        data={
          'year':str(y),'displayYear':y,
          'title':{'year':str(y),'displayYear':y,'ganzhi':s(yr.get('干支')),
                   'monthCount':expected_months,'dayCount':expected,
                   'yuanriGanzhi':s(yr.get('年初干支')),'solarNewYear':s(yr.get('年初西曆')),
                   'winterSolstice':s(yr.get('冬至')),'eras':groups},
          'months':months
        }
        dump_js(ROOT/'data'/f'{y}.js',data)
        generated[y]=data

    # Update chronology year cards from year page only.
    pidx,obj=load_index()
    year_map={int(x['year']):x for x in obj['years']}
    for y,data in generated.items():
        rec=year_map[y]
        details=[]; eras=[]; states=[]
        for group in data['title']['eras']:
            state=group['state']
            if state not in states: states.append(state)
            for item in group['entries']:
                era=item.get('era',''); note=item.get('source',''); notes=item.get('notes',[])
                eras.append(era)
                details.append({'state':state,'text':era,'era':era,'note':note,'notes':notes,
                                'fullText':(era+(' '+note if note else ''))})
        solar=data['title']['solarNewYear']
        rec.update({'ganzhi':data['title']['ganzhi'],'solarYear':solar.split('-')[0],
                    'solarNewYear':solar,'eras':eras,'eraDetails':details,'states':states})

    # Rebuild only the Song dynasty card/ruler filter from the newly rebuilt daily records,
    # preserving 1180 from the adjacent live-derived file already in the reference package.
    song=next(d for d in obj['dynasties'] if d.get('id')=='宋')
    song_years=[]; ruler_years=defaultdict(set)
    for y in range(1180,1501):
        if y in generated:
            data=generated[y]
        else:
            p=ROOT/'data'/f'{y}.js'; text=p.read_text(encoding='utf-8')
            m=re.search(r'window\.CalendarData\s*=\s*(\{.*?\});\s*\n\nwindow\.CalendarData\.dayIndex',text,re.S)
            data=json.loads(m.group(1))
        hit=False
        for mo in data.get('months',[]):
            for day in mo.get('days',[]):
                for r in day.get('orthodoxies',[]):
                    if r.get('state')=='宋':
                        hit=True
                        if r.get('ruler'): ruler_years[r['ruler']].add(y)
        if hit: song_years.append(y)
    song['states']=['宋']; song['years']=song_years
    song['displayRange']=f'{song_years[0]:04d}年至{song_years[-1]:04d}年'
    song['displayCount']=f'凡{len(song_years)}年'
    rulers=[]
    for name,ys in ruler_years.items():
        sy=sorted(ys); rulers.append({'name':name,'years':sy,'range':f'{sy[0]:04d}年至{sy[-1]:04d}年','count':len(sy)})
    rulers.sort(key=lambda x:(x['years'][0],x['name']))
    song['rulers']=rulers; song['rulerCount']=len(rulers); song['allRulersLabel']=f'全部國君（{len(rulers)}）'
    pidx.write_text('window.CalendarIndexData = '+json.dumps(obj,ensure_ascii=False,separators=(',',':'))+';\n',encoding='utf-8')

    # Retain a reproducible script and a data-source snapshot inside the package.
    print(json.dumps({'years':len(generated),'days':total,'songYears':len(song_years),
                      'songRange':[song_years[0],song_years[-1]],'songRulers':len(rulers)},ensure_ascii=False))

if __name__=='__main__': main()
